import openpyxl

wb = openpyxl.load_workbook("c:\\Users\\vatty\\Desktop\\TestData.xlsx")
print(wb.sheetnames)

print("Active Sheet--->"+wb.active.title)

sh = wb['Sheet1']

maxrows = sh.max_row
maxcolms = sh.max_column

print(sh.title)
print(sh['A4'].value+' '+sh['B4'].value)

print(f'Total Rows {maxrows} and Columns {maxcolms}')


for r in range(1, maxrows+1):
    for c in range(1, maxcolms+1):
        c = sh.cell(r, c)
        print(c.value)