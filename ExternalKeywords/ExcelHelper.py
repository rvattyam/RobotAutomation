import openpyxl


wb = openpyxl.load_workbook("C:\\Users\\vatty\\Desktop\\TestData.xlsx")

sh = wb['Sheet1']

for i in range(2,7):
    cellp = sh.cell(i,2)
    cellu = sh.cell(i,1)
    print(cellu.value,cellp.value)